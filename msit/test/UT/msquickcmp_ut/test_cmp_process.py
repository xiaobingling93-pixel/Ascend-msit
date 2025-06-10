# Copyright (c) 2023-2024 Huawei Technologies Co., Ltd.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os, sys
import shutil
import unittest
import tempfile
from unittest.mock import Mock, MagicMock
from unittest.mock import patch, MagicMock, mock_open

import pandas as pd
import openpyxl
import pytest

from auto_optimizer.graph_refactor import Node
from auto_optimizer import OnnxGraph
from components.debug.compare.msquickcmp.cmp_process import (
    _generate_golden_data_model,
    _correct_the_wrong_order,
    _check_output_node_name_mapping,
    _get_single_csv_in_folder,
    _read_and_process_csv,
    _write_csv,
    AccuracyCompareException,
    mindir_to_om_process, check_dump_and_compare, cmp_process,
    dump_and_compare, compare_run, compare_process, run_om_model_compare, run,
    _append_column_to_csv, run, print_advisor_info, CmpArgsAdapter, fusion_close_model_convert, single_op_compare, \
    find_accuracy_interval
)


class TestGenerateGoldenDataModel(unittest.TestCase):
    @patch('components.debug.compare.msquickcmp.cmp_process.is_saved_model_valid', return_value=True)
    @patch('msquickcmp.tf.tf_save_model_dump_data.TfSaveModelDumpData')
    def test_generate_tf_saved_model(self, mock_tf_class, mock_is_valid):
        args = MagicMock()
        args.model_path = 'model_path'
        instance = MagicMock()
        mock_tf_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'dump_path')
        self.assertEqual(result, instance)
        self.assertIsNone(ext)

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.prototxt'))
    @patch('msquickcmp.caffe_model.caffe_dump_data.CaffeDumpData')
    def test_generate_caffe_model(self, mock_caffe_class, _):
        args = MagicMock()
        args.model_path = 'model.prototxt'
        args.weight_path = 'weight'
        instance = MagicMock()
        mock_caffe_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'dump_path')
        self.assertEqual(result, instance)
        self.assertEqual(ext, '.prototxt')

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.pb'))
    @patch('msquickcmp.tf.tf_dump_data.TfDumpData')
    def test_generate_tf_pb_model(self, mock_tf_class, _):
        args = MagicMock()
        args.model_path = 'model.pb'
        instance = MagicMock()
        mock_tf_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'dump_path')
        self.assertEqual(result, instance)
        self.assertEqual(ext, '.pb')

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.onnx'))
    @patch('msquickcmp.onnx_model.onnx_dump_data.OnnxDumpData')
    def test_generate_onnx_model(self, mock_onnx_class, _):
        args = MagicMock()
        args.model_path = 'model.onnx'
        instance = MagicMock()
        mock_onnx_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'npu_path')
        self.assertEqual(result, instance)
        self.assertEqual(ext, '.onnx')

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.om'))
    @patch('components.debug.compare.msquickcmp.cmp_process.NpuDumpData')
    def test_generate_om_model(self, mock_npu_class, _):
        args = MagicMock()
        args.model_path = 'model.om'
        instance = MagicMock()
        mock_npu_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'npu_path')
        self.assertEqual(result, instance)
        self.assertEqual(ext, '.om')

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.unsupported'))
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    def test_generate_unsupported_model(self, mock_logger, _):
        args = MagicMock()
        args.model_path = 'model.unsupported'

        with self.assertRaises(AccuracyCompareException):
            _generate_golden_data_model(args, 'npu_path')


class TestCorrectTheWrongOrder(unittest.TestCase):
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    def test_correct_order(self, mock_logger):
        data = {0: 'a', 1: 'b'}
        _correct_the_wrong_order(0, 1, data)
        self.assertEqual(data[0], 'b')
        self.assertEqual(data[1], 'a')


class TestCheckOutputNodeNameMapping(unittest.TestCase):
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    def test_check_output_node_name_mapping_match(self, mock_logger):
        output_node = {0: 'abc/xyz:0'}
        golden_info = {0: 'abc_xyz.0.out'}
        _check_output_node_name_mapping(output_node, golden_info)

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger.warning')
    def test_check_output_node_name_mapping_no_match(self, mock_warn):
        output_node = {0: 'notmatch'}
        golden_info = {1: 'no_match_file.out'}
        _check_output_node_name_mapping(output_node, golden_info)
        mock_warn.assert_called()


class TestGetSingleCsvInFolder(unittest.TestCase):
    @patch('os.listdir', return_value=['data.csv'])
    def test_get_csv_found(self, _):
        path = '/some/path'
        result = _get_single_csv_in_folder(path)
        self.assertEqual(result, '/some/path/data.csv')

    @patch('os.listdir', return_value=['no_csv.txt'])
    def test_get_csv_not_found(self, _):
        with self.assertRaises(IOError):
            _get_single_csv_in_folder('/some/dir')


class TestReadAndProcessCSV(unittest.TestCase):
    @patch('components.debug.compare.msquickcmp.cmp_process.ms_open', new_callable=mock_open, read_data='a,b\n1,2')
    @patch('components.debug.compare.msquickcmp.cmp_process.Rule.input_file')
    def test_read_and_process_csv(self, mock_rule, mock_file):
        mock_rule.return_value.check.return_value = True

        process_func = MagicMock(return_value=[['header1', 'header2'], ['val1', 'val2']])
        result = _read_and_process_csv('fake.csv', process_func, ['node'])
        self.assertTrue(isinstance(result, list))
        self.assertEqual(result[0], ['header1', 'header2'])


class TestWriteCsv(unittest.TestCase):
    @patch('components.debug.compare.msquickcmp.cmp_process.ms_open', new_callable=mock_open)
    @patch('components.debug.compare.msquickcmp.cmp_process.csv.writer')
    @patch('components.debug.compare.msquickcmp.cmp_process.sanitize_csv_value', side_effect=lambda x: x)
    def test_write_csv(self, mock_sanitize, mock_csv_writer, mock_file):
        writer = MagicMock()
        mock_csv_writer.return_value = writer

        _write_csv('path.csv', [['a', 'b'], ['1', '2']])
        self.assertEqual(writer.writerows.call_count, 1)



@pytest.fixture(scope="function")
def import_cmp_process():
    backup = {}
    for mod in ['acl', 'msquickcmp.cmp_process']:
        if mod in sys.modules:
            backup[mod] = sys.modules[mod]
    mock_acl = MagicMock()
    sys.modules['acl'] = mock_acl
    from msquickcmp.cmp_process import csv_sum, _get_model_output_node_name_list, \
        _find_previous_node
    functions = {
        "csv_sum": csv_sum,
        "_get_model_output_node_name_list": _get_model_output_node_name_list,
        "_find_previous_node": _find_previous_node
    }
    yield functions
    
    for mod, module_obj in backup.items():
        sys.modules[mod] = module_obj
    for mod in ['acl', 'msquickcmp.cmp_process']:
        if mod not in backup and mod in sys.modules:
            del sys.modules[mod]


@pytest.fixture(scope="function")
def generate_fake_path():
    cur_dir = os.path.dirname(os.path.realpath(__file__))
    resource_dir = os.path.join(cur_dir, 'test_csv_sum', "2023072009")

    os.makedirs(resource_dir, 0o750, exist_ok=True)
    os.makedirs(os.path.join(resource_dir, "images-2_3_638_640"), 0o750, exist_ok=True)
    os.makedirs(os.path.join(resource_dir, "images-2_3_640_640"), 0o750, exist_ok=True)

    df1 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df1.to_csv(os.path.join(resource_dir, "images-2_3_638_640", "file1.csv"), index=False)

    df2 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df2.to_csv(os.path.join(resource_dir, "images-2_3_640_640", "file2.csv"), index=False)

    with pd.ExcelWriter(os.path.join(cur_dir, "test_csv_sum", "expected_output.xlsx")) as writer:
        df1.to_excel(writer, sheet_name='images-2_3_638_640', index=False)
        df2.to_excel(writer, sheet_name='images-2_3_640_640', index=False)
    yield resource_dir

    shutil.rmtree(os.path.join(cur_dir, "test_csv_sum"))


def test_csv_sum_given_path_when_valid_then_pass(import_cmp_process, generate_fake_path):
    csv_sum = import_cmp_process["csv_sum"]
    csv_sum(generate_fake_path)
    expected_output = openpyxl.load_workbook(os.path.join(generate_fake_path, '..', 'expected_output.xlsx'))
    result_summary = openpyxl.load_workbook(os.path.join(generate_fake_path, 'result_summary.xlsx'))

    sheets1 = result_summary.sheetnames
    sheets2 = expected_output.sheetnames

    assert len(sheets1) == len(sheets2)

    for sheet_name in sheets1:
        sheet1 = result_summary[sheet_name]
        sheet2 = expected_output[sheet_name]

        assert sheet1.max_row == sheet2.max_row
        assert sheet1.max_column == sheet2.max_column

        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                assert sheet1.cell(row=row, column=col).value == sheet2.cell(row=row, column=col).value


class TestProcessIsNpuAndIsPrecisionErrorOps(unittest.TestCase):
    def setUp(self):
        self.backup = {}
        for mod in ['acl', 'msquickcmp.cmp_process']:
            if mod in sys.modules:
                self.backup[mod] = sys.modules[mod]
        mock_acl = MagicMock()
        sys.modules['acl'] = mock_acl
        from msquickcmp.cmp_process import _process_is_npu_and_is_precision_error_ops
        self._process_is_npu_and_is_precision_error_ops = _process_is_npu_and_is_precision_error_ops

    def tearDown(self):
        for mod, module_obj in self.backup.items():
            sys.modules[mod] = module_obj
        for mod in ['acl', 'msquickcmp.cmp_process']:
            if mod not in self.backup and mod in sys.modules:
                del sys.modules[mod]


    def test_process_function(self):
        self.header = ["GroundTruth", "OpType", "CosineSimilarity", 
                       "RelativeEuclideanDistance", "KullbackLeiblerDivergence", 
                       "RootMeanSquareError", "MeanRelativeError"]
        self.rows = [
            [],
            ["*", "Add", "0.85", "0.12", "0.05", "0.04", "0.03"],  # NPU op, precision error
            ["*", "Add", "0.999", "0.005", "0.001", "0.04", "0.03"], # NPU op, no precision error
            # not NPU op, precision error
            ["/extractor/Softmax", "Add", "0.85", "0.12", "0.05", "0.04", "0.03"],  
            # not NPU op, no precision error
            ["/extractor/Softmax", "Add", "0.999", "0.005", "0.001", "0.04", "0.03"],
            # RelativeEuclideanDistance 
            ["/extractor/Softmax", "Add", "0.999", "0.06", "0.001", "0.04", "0.03"],
            # KullbackLeiblerDivergence
            ["/extractor/Reshape_3", "Add", "0.999", "0.005", "0.007", "0.04", "0.03"],
            # RootMeanSquareError
            ["/extractor/Softmax", "Add", "0.999", "0.005", "0.001", "1.1", "0.03"], 
            # MeanRelativeError
            ["/extractor/Concat_1", "Add", "0.999", "0.005", "0.001", "0.04", "10"] 
        ]
        node_output_name_list = ["/extractor/Reshape_3", "/extractor/Concat_1"]
        
        processed_rows = self._process_is_npu_and_is_precision_error_ops(self.header, self.rows, node_output_name_list)

        self.assertIn('IsNpuOps', self.header)
        self.assertIn('IsOutputNode', self.header)
        self.assertIn('IsPrecisionError', self.header)

        expected_results = [
            ["YES", "NO", "YES"],  
            ["YES", "NO", "NO"],  
            ["NO", "NO", "YES"],  
            ["NO", "NO", "NO"],
            ["NO", "NO", "YES"],
            ["NO", "YES", "YES"],
            ["NO", "NO", "YES"],
            ["NO", "YES", "YES"]     
        ]
        
        for i, row in enumerate(processed_rows[1:]):
            self.assertEqual(row[-3:], expected_results[i])


class TestIsRowPrecisionError(unittest.TestCase):
    def setUp(self):
        self.backup = {}
        for mod in ['acl', 'msquickcmp.cmp_process']:
            if mod in sys.modules:
                self.backup[mod] = sys.modules[mod]
        mock_acl = MagicMock()
        sys.modules['acl'] = mock_acl
        from msquickcmp.cmp_process import _is_row_precision_error
        self._is_row_precision_error = _is_row_precision_error

    def tearDown(self):
        for mod, module_obj in self.backup.items():
            sys.modules[mod] = module_obj
        for mod in ['acl', 'msquickcmp.cmp_process']:
            if mod not in self.backup and mod in sys.modules:
                del sys.modules[mod]

    def test_below_thresholds(self):
        # All metrics are below the threshold
        result = self._is_row_precision_error(0.95, 0.05, 0.04, 0.01, 0.005)
        self.assertTrue(result)

    def test_cosine_similarity_below_threshold(self):
        # only cosine similarity is below the threshold
        result = self._is_row_precision_error(0.85, 0.05, 0.04, 0.01, 0.005)
        self.assertTrue(result)

    def test_relative_euclidean_distance_above_threshold(self):
        # only relative Euclidean distance above threshold
        result = self._is_row_precision_error(0.95, 0.15, 0.04, 0.01, 0.005)
        self.assertTrue(result)

    def test_kullback_leibler_divergence_above_threshold(self):
        # only Kullback-Leibler above threshold
        result = self._is_row_precision_error(0.95, 0.05, 0.06, 0.01, 0.005)
        self.assertTrue(result)

    def test_root_mean_square_error_above_threshold(self):
        # only RootMeanSquareError is above the threshold
        result = self._is_row_precision_error(0.95, 0.05, 0.04, 0.03, 0.005)
        self.assertTrue(result)

    def test_mean_relative_error_above_threshold(self):
        # only mean relative error is above the threshold
        result = self._is_row_precision_error(0.95, 0.05, 0.04, 0.01, 0.02)
        self.assertTrue(result)

    def test_multiple_errors(self):
        # Multiple indicators do not meet the requirements
        result = self._is_row_precision_error(0.85, 0.15, 0.06, 0.03, 0.02)
        self.assertTrue(result)
        
        
class TestIsOutputNodeError(unittest.TestCase):
    def setUp(self):
        self.backup = {}
        for mod in ['acl', 'msquickcmp.cmp_process']:
            if mod in sys.modules:
                self.backup[mod] = sys.modules[mod]
        mock_acl = MagicMock()
        sys.modules['acl'] = mock_acl
        from msquickcmp.cmp_process import _is_output_node
        self._is_output_node = _is_output_node

    def tearDown(self):
        for mod, module_obj in self.backup.items():
            sys.modules[mod] = module_obj
        for mod in ['acl', 'msquickcmp.cmp_process']:
            if mod not in self.backup and mod in sys.modules:
                del sys.modules[mod]
    
    def test__is_output_node(self):
        node_output_name_list = []
        result_false = self._is_output_node("/extractor/Concat_1", node_output_name_list)
        self.assertFalse(result_false)
        
        node_output_name_list = ["/extractor/Reshape_3", "/extractor/Concat_1"]
        result_true = self._is_output_node("/extractor/Reshape_3", node_output_name_list)
        self.assertTrue(result_true)
        
        result_false = self._is_output_node("", node_output_name_list)
        self.assertFalse(result_false)


class MockSession:
    def get_outputs(self):
        return [MockNode(name="output_1", outputs=["net_output"]), MockNode(name="output_2", outputs=["net_output"])]


class MockGraph:
    def __init__(self, nodes):
        self.node = nodes


class MockNode:
    def __init__(self, name, outputs):
        self.name = name
        self.output = outputs


def test_find_previous_node(import_cmp_process):
    _find_previous_node = import_cmp_process["_find_previous_node"]
    # 创建模拟的节点
    node1 = MockNode(name="node1", outputs=["output_1"])
    node2 = MockNode(name="node2", outputs=["output_2"])
    graph = MockGraph(nodes=[node1, node2])

    # 测试找到前一个节点
    assert _find_previous_node(graph, "output_1") == "node1"
    assert _find_previous_node(graph, "output_2") == "node2"

    # 测试未找到前一个节点
    assert _find_previous_node(graph, "output_3") is None


def test_get_model_output_node_name_list(import_cmp_process):
    _get_model_output_node_name_list = import_cmp_process["_get_model_output_node_name_list"]
    # 创建模拟的会话和模型
    session = MockSession()
    node1 = MockNode(name="node1", outputs=["output_1"])
    node2 = MockNode(name="node2", outputs=["output_2"])
    origin_model = Mock(graph=MockGraph(nodes=[node1, node2]))
    # 测试正常情况
    result = _get_model_output_node_name_list(session, origin_model)
    assert result == ["node1", "node2"], f"Expected ['node1', 'node2'], but got {result}"

    # 测试找不到前一个节点的情况
    node3 = MockNode(name="node3", outputs=["output_3"])
    origin_model.graph.node = [node3]  # 修改模型节点
    result = _get_model_output_node_name_list(session, origin_model)
    assert result is None, f"Expected None, but got {result}"


class TestDumpCmpFunctions(unittest.TestCase):
    @patch("components.debug.compare.msquickcmp.cmp_process.utils.logger")
    @patch("components.debug.compare.msquickcmp.cmp_process.os.path.exists", return_value=True)
    @patch("components.debug.compare.msquickcmp.cmp_process.site.getsitepackages", return_value=["/mock/site"])
    @patch("components.debug.compare.msquickcmp.cmp_process.shutil.move")
    @patch("components.debug.compare.msquickcmp.cmp_process.utils.execute_command")
    @patch("components.debug.compare.msquickcmp.cmp_process.os.path.realpath", return_value="/real/path")
    @patch("components.debug.compare.msquickcmp.cmp_process.check_write_directory")
    def test_mindir_to_om_process(self, mock_check, mock_realpath, mock_exec, mock_move,
                                   mock_site, mock_exists, mock_logger):
        args = MagicMock()
        args.model_path = "model.onnx"
        args.offline_model_path = "model.mindir"
        os.environ.pop("LD_PRELOAD", None)
        result = mindir_to_om_process(args)
        self.assertTrue(result)
        self.assertIn("LD_PRELOAD", os.environ)

    def test_check_dump_and_compare_all_none(self):
        args = MagicMock(my_path=None, golden_path=None, ops_json=None)
        self.assertTrue(check_dump_and_compare(args))

    def test_check_dump_and_compare_valid(self):
        args = MagicMock(my_path="1", golden_path="2", ops_json="3")
        self.assertFalse(check_dump_and_compare(args))

    def test_check_dump_and_compare_invalid(self):
        args = MagicMock(my_path="1", golden_path=None, ops_json="3")
        with self.assertRaises(Exception):
            check_dump_and_compare(args)

    @patch("components.debug.compare.msquickcmp.cmp_process.dump_and_compare")
    @patch("components.debug.compare.msquickcmp.cmp_process.compare_run")
    @patch("components.debug.compare.msquickcmp.cmp_process.os.path.realpath", return_value="/real/path")
    def test_cmp_process_dump(self, mock_realpath, mock_compare, mock_dump):
        args = MagicMock()
        args.my_path = None
        args.golden_path = None
        args.ops_json = None
        cmp_process(args, use_cli=True)
        mock_dump.assert_called_once()

    @patch("components.debug.compare.msquickcmp.cmp_process.compare_process", return_value=[{"GroundTruth": "node1"}])
    @patch("components.debug.compare.msquickcmp.cmp_process.find_accuracy_interval", return_value=[["error1", "error2"]])
    @patch("components.debug.compare.msquickcmp.cmp_process.ms_open", new_callable=mock_open)
    @patch("components.debug.compare.msquickcmp.cmp_process.os.path.join", return_value="/path/info.txt")
    def test_compare_run(self, mock_join, mock_open_file, mock_find, mock_compare_process):
        args = MagicMock()
        args.locat = True
        args.out_path = "/out"
        compare_run(args)
        mock_compare_process.assert_called_once()

    @patch("components.debug.compare.msquickcmp.cmp_process.NetCompare")
    @patch("components.debug.compare.msquickcmp.cmp_process.analyser.Analyser", return_value=MagicMock(return_value=([], [])))
    @patch("components.debug.compare.msquickcmp.cmp_process.print_advisor_info")
    @patch("components.debug.compare.msquickcmp.cmp_process._append_column_to_csv")
    def test_compare_process(self, mock_append, mock_print, mock_analyser, mock_netcompare):
        args = MagicMock()
        args.my_path = "mypath"
        args.golden_path = "golden"
        args.ops_json = "ops.json"
        args.locat = False
        result = compare_process(args)
        self.assertEqual(result, [])

    @patch("components.debug.compare.msquickcmp.cmp_process.run_om_model_compare", return_value=["dummy"])
    @patch("components.debug.compare.msquickcmp.cmp_process.get_shape_to_directory_name", return_value="shapedir")
    @patch("components.debug.compare.msquickcmp.cmp_process.os.path.join", return_value="/joined/path")
    @patch("components.debug.compare.msquickcmp.cmp_process.is_saved_model_valid", return_value=False)
    def test_run_invalid_model(self, mock_valid, mock_join, mock_shape, mock_run_om):
        args = MagicMock()
        args.model_path = "model.om"
        args.offline_model_path = "off.om"
        args.out_path = "/out"
        result = run(args, input_shape="1,2,3", original_out_path="/base", use_cli=False)
        self.assertEqual(result, ["dummy"])


class TestAppendColumnToCSV(unittest.TestCase):

    @patch("components.debug.compare.msquickcmp.cmp_process._write_csv")
    @patch("components.debug.compare.msquickcmp.cmp_process._read_and_process_csv")
    @patch("components.debug.compare.msquickcmp.cmp_process._get_single_csv_in_folder")
    @patch("components.debug.compare.msquickcmp.cmp_process._process_is_npu_and_is_precision_error_ops")
    def test_append_column_to_csv_with_node_output_show_list(self, mock_process_func, mock_get_csv, mock_read_process, mock_write):
        mock_get_csv.return_value = "/fake/path/file.csv"
        mock_read_process.return_value = [["row1"], ["row2"]]
        node_output_show_list = ["node1", "node2"]
        _append_column_to_csv("/some/path", node_output_show_list)
        mock_get_csv.assert_called_once_with("/some/path")
        mock_read_process.assert_called_once_with("/fake/path/file.csv", mock_process_func, node_output_show_list)
        mock_write.assert_called_once_with("/fake/path/file.csv", [["row1"], ["row2"]])

    @patch("components.debug.compare.msquickcmp.cmp_process._write_csv")
    @patch("components.debug.compare.msquickcmp.cmp_process._read_and_process_csv")
    @patch("components.debug.compare.msquickcmp.cmp_process._get_single_csv_in_folder")
    @patch("components.debug.compare.msquickcmp.cmp_process._process_is_npu_and_is_precision_error_ops")
    def test_append_column_to_csv_with_none_node_output_show_list(self, mock_process_func, mock_get_csv, mock_read_process, mock_write):
        mock_get_csv.return_value = "/fake/path/file.csv"
        mock_read_process.return_value = [["row1"]]
        _append_column_to_csv("/some/path", None)
        mock_get_csv.assert_called_once_with("/some/path")
        mock_read_process.assert_called_once_with("/fake/path/file.csv", mock_process_func, [])
        mock_write.assert_called_once_with("/fake/path/file.csv", [["row1"]])


class TestRunFunction(unittest.TestCase):
    def test_run_valid_model(self):
        import types
        fake_tf_common = types.ModuleType("tf_common")
        fake_tf_common.some_function = MagicMock()

        with patch.dict(sys.modules, {
            "msquickcmp.common.tf_common": fake_tf_common,
            "msquickcmp.tf.tf_save_model_dump_data": MagicMock(),
            "msquickcmp.npu.npu_tf_adapter_dump_data": MagicMock(),
        }), \
        patch("components.debug.compare.msquickcmp.cmp_process._append_column_to_csv"), \
        patch("components.debug.compare.msquickcmp.cmp_process.print_advisor_info"), \
        patch("components.debug.compare.msquickcmp.cmp_process.analyser.Analyser") as mock_analyser_cls, \
        patch("components.debug.compare.msquickcmp.cmp_process.NetCompare") as mock_net_compare_cls, \
        patch("components.debug.compare.msquickcmp.cmp_process.is_saved_model_valid") as mock_is_valid, \
        patch("components.debug.compare.msquickcmp.cmp_process.get_shape_to_directory_name") as mock_get_dir, \
        patch("components.debug.compare.msquickcmp.cmp_process.utils"):
            from components.debug.compare.msquickcmp import cmp_process
            args = MagicMock()
            args.input_shape = None
            args.out_path = "some_path"
            args.offline_model_path = "offline"
            args.model_path = "model"
            args.locat = False
            mock_is_valid.return_value = True
            mock_get_dir.return_value = "shapedir"
            npu_mock_instance = MagicMock()
            npu_mock_instance.generate_dump_data.return_value = ("npu_path", "json_path")
            tf_mock_instance = MagicMock()
            tf_mock_instance.generate_dump_data.return_value = "golden_path"
            sys.modules["msquickcmp.npu.npu_tf_adapter_dump_data"].NpuTfAdapterDumpData.return_value = npu_mock_instance
            sys.modules["msquickcmp.tf.tf_save_model_dump_data"].TfSaveModelDumpData.return_value = tf_mock_instance
            net_compare = MagicMock()
            mock_net_compare_cls.return_value = net_compare
            mock_analyser = MagicMock()
            mock_analyser.return_value = (["invalid_row"], None)
            mock_analyser_cls.return_value = mock_analyser
            result = cmp_process.run(args, [1, 3, 224, 224], "output", use_cli=True)
            self.assertEqual(result, ["invalid_row"])

    @patch("components.debug.compare.msquickcmp.cmp_process.run_om_model_compare")
    @patch("components.debug.compare.msquickcmp.cmp_process.is_saved_model_valid")
    def test_run_with_invalid_saved_model(self, mock_is_valid, mock_run_om):
        mock_args = MagicMock()
        mock_args.out_path = "some_path"
        mock_is_valid.return_value = False
        mock_run_om.return_value = ["rowA", "rowB"]
        from components.debug.compare.msquickcmp.cmp_process import run
        result = run(mock_args, input_shape=None, original_out_path="orig", use_cli=False)
        self.assertEqual(result, ["rowA", "rowB"])
        mock_run_om.assert_called_once_with(mock_args, False)


class TestRunOmModelCompare(unittest.TestCase):

    @patch('components.debug.compare.msquickcmp.cmp_process.print_advisor_info')
    @patch('components.debug.compare.msquickcmp.cmp_process._append_column_to_csv')
    @patch('components.debug.compare.msquickcmp.cmp_process._get_model_output_node_name_list')
    @patch('components.debug.compare.msquickcmp.cmp_process.analyser.Analyser')
    @patch('components.debug.compare.msquickcmp.cmp_process._check_output_node_name_mapping')
    @patch('components.debug.compare.msquickcmp.cmp_process.NetCompare')
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.handle_ground_truth_files')
    @patch('components.debug.compare.msquickcmp.cmp_process.is_saved_model_valid')
    @patch('components.debug.compare.msquickcmp.cmp_process._generate_golden_data_model')
    @patch('components.debug.compare.msquickcmp.cmp_process.convert_bin_dump_data_to_npy')
    @patch('components.debug.compare.msquickcmp.cmp_process.NpuDumpData')
    @patch('components.debug.compare.msquickcmp.cmp_process.OmParser')
    @patch('components.debug.compare.msquickcmp.cmp_process.atc_utils.convert_model_to_json')
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    def test_run_om_model_compare_basic(self, mock_logger, mock_convert_model_to_json, mock_OmParser, 
                                       mock_NpuDumpData, mock_convert_bin_dump_data_to_npy, 
                                       mock_generate_golden_data_model, mock_is_saved_model_valid,
                                       mock_handle_ground_truth_files, mock_NetCompare,
                                       mock_check_output_node_name_mapping, mock_Analyser,
                                       mock_get_model_output_node_name_list, mock_append_column_to_csv,
                                       mock_print_advisor_info):

        args = MagicMock()
        args.cann_path = "/fake/cann"
        args.offline_model_path = "/fake/offline.om"
        args.out_path = "/fake/out"
        args.model_path = "/fake/model.om"
        args.fusion_switch_file = None
        args.bin2npy = False
        args.custom_op = ""
        args.dump = False
        args.locat = False

        mock_convert_model_to_json.side_effect = lambda cann, model, out: "/fake/json_path"

        omparser_instance = MagicMock()
        omparser_instance.get_aipp_config_content.return_value = True
        mock_OmParser.return_value = omparser_instance

        npu_dump_instance = MagicMock()
        mock_NpuDumpData.return_value = npu_dump_instance

        npu_dump_instance.generate_inputs_data.return_value = None

        npu_dump_instance.generate_dump_data.return_value = ("/fake/npu_dump_path", "/fake/npu_net_output_path")

        npu_dump_instance.get_expect_output_name.return_value = ["output_node"]

        mock_convert_bin_dump_data_to_npy.return_value = ""

        golden_dump_mock = MagicMock()
        mock_generate_golden_data_model.return_value = (golden_dump_mock, ".onnx")

        mock_is_saved_model_valid.return_value = True
        golden_dump_mock.generate_inputs_data.return_value = None
        golden_dump_mock.generate_dump_data.return_value = "/fake/golden_dump_path"
        golden_dump_mock.get_net_output_info.return_value = {"some": "info"}

        mock_handle_ground_truth_files.return_value = None

        net_compare_instance = MagicMock()
        mock_NetCompare.return_value = net_compare_instance
        net_compare_instance.net_output_compare.return_value = None

        mock_check_output_node_name_mapping.return_value = None

        analyser_instance = MagicMock()
        analyser_instance.return_value = ("invalid_rows_mock", "other")
        mock_Analyser.return_value = analyser_instance

        mock_get_model_output_node_name_list.return_value = ["node1", "node2"]

        mock_print_advisor_info.return_value = None
        mock_append_column_to_csv.return_value = None

        result = run_om_model_compare(args, use_cli=True)

        self.assertEqual(result, "invalid_rows_mock")

        mock_convert_model_to_json.assert_called()
        mock_OmParser.assert_called()
        npu_dump_instance.generate_inputs_data.assert_called()
        npu_dump_instance.generate_dump_data.assert_called()
        mock_generate_golden_data_model.assert_called()
        mock_is_saved_model_valid.assert_called_with(args.model_path)
        golden_dump_mock.generate_inputs_data.assert_called_with(True)
        golden_dump_mock.generate_dump_data.assert_called_with("/fake/json_path")
        mock_handle_ground_truth_files.assert_called()
        mock_NetCompare.assert_called()
        net_compare_instance.net_output_compare.assert_called()
        mock_check_output_node_name_mapping.assert_called()
        mock_Analyser.assert_called()
        mock_get_model_output_node_name_list.assert_called()
        mock_print_advisor_info.assert_called_with(args.out_path)
        mock_append_column_to_csv.assert_called()

    @patch('components.debug.compare.msquickcmp.cmp_process.OmParser')
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    def test_raise_when_fusion_switch_file_and_aipp(self, mock_logger, mock_OmParser):
        args = MagicMock()
        args.cann_path = "a"
        args.offline_model_path = "b"
        args.out_path = "c"
        args.model_path = "model.om"
        args.fusion_switch_file = "fusion_switch"
        args.bin2npy = False
        args.custom_op = ""
        args.dump = False
        args.locat = False

        omparser_instance = MagicMock()
        omparser_instance.get_aipp_config_content.return_value = True
        mock_OmParser.return_value = omparser_instance
        with self.assertRaises(AccuracyCompareException):
            run_om_model_compare(args, use_cli=True)
        mock_logger.error.assert_called()


class TestPrintAdvisorInfo(unittest.TestCase):
    
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    @patch('components.debug.compare.msquickcmp.cmp_process.ms_open', create=True)
    @patch('components.debug.compare.msquickcmp.cmp_process.load_file_to_read_common_check')
    @patch('components.debug.compare.msquickcmp.cmp_process.Rule.input_file')
    @patch('components.debug.compare.msquickcmp.cmp_process.os.path.join')
    def test_print_advisor_info_file_exists(self, mock_os_path_join, mock_rule_input_file,
                                            mock_load_file_to_read_common_check, mock_ms_open, mock_logger):

        mock_os_path_join.return_value = '/fake/path/advisor_summary.txt'
        mock_check = MagicMock()
        mock_rule_input_file.return_value.check = mock_check
        mock_check.return_value = True

        mock_load_file_to_read_common_check.return_value = '/fake/path/advisor_summary.txt'
        mock_file = MagicMock()
        mock_file.readlines.return_value = ['line1\n', 'line2\n']
        mock_ms_open.return_value.__enter__.return_value = mock_file

        print_advisor_info('/fake/path')

        mock_os_path_join.assert_called_once_with('/fake/path', 'advisor_summary.txt')
        mock_check.assert_called_once_with('/fake/path/advisor_summary.txt')
        mock_load_file_to_read_common_check.assert_called_once_with('/fake/path/advisor_summary.txt')
        mock_ms_open.assert_called_once_with('/fake/path/advisor_summary.txt', 'r', max_size=10 * 1024 * 1024 * 1024)
        mock_file.readlines.assert_called_once()

        mock_logger.info.assert_any_call('The advisor summary (.txt) is saved in :"/fake/path/advisor_summary.txt"')
        mock_logger.info.assert_any_call('line1')
        mock_logger.info.assert_any_call('line2')

    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    @patch('components.debug.compare.msquickcmp.cmp_process.Rule.input_file')
    @patch('components.debug.compare.msquickcmp.cmp_process.os.path.join')
    def test_print_advisor_info_file_not_exists(self, mock_os_path_join, mock_rule_input_file, mock_logger):
        # Arrange
        mock_os_path_join.return_value = '/fake/path/advisor_summary.txt'

        mock_check = MagicMock()
        mock_check.return_value = False
        mock_rule_input_file.return_value.check = mock_check

        # Act
        print_advisor_info('/fake/path')

        # Assert
        mock_os_path_join.assert_called_once_with('/fake/path', 'advisor_summary.txt')
        mock_check.assert_called_once_with('/fake/path/advisor_summary.txt')
        mock_logger.info.assert_not_called()


class TestFusionCloseModelConvert(unittest.TestCase):
    def test_fusion_close_model_convert_no_fusion_file(self):
        args = CmpArgsAdapter(gold_model='some_gold_model', om_model='some_om_model')
        args.fusion_switch_file = None

        with patch('components.debug.compare.msquickcmp.cmp_process.utils.execute_command') as mock_exec_cmd:
            fusion_close_model_convert(args)
            mock_exec_cmd.assert_not_called()


class TestSingleOpCompare(unittest.TestCase):
    @patch('components.debug.compare.msquickcmp.cmp_process.sp.atc_conversion')
    @patch('components.debug.compare.msquickcmp.cmp_process.sp.find_all_csv')
    @patch('components.debug.compare.msquickcmp.cmp_process.run')
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.merge_csv')
    @patch('components.debug.compare.msquickcmp.cmp_process.shutil.rmtree')
    @patch('components.debug.compare.msquickcmp.cmp_process.time.strftime')
    @patch('components.debug.compare.msquickcmp.cmp_process.ms_makedirs')
    @patch('components.debug.compare.msquickcmp.cmp_process.al.create_bin_file')
    @patch('components.debug.compare.msquickcmp.cmp_process.al.find_npy_files_with_prefix')
    @patch('components.debug.compare.msquickcmp.cmp_process.al.input_completion')
    @patch('components.debug.compare.msquickcmp.cmp_process.Rule')
    @patch('components.debug.compare.msquickcmp.cmp_process.onnxruntime.InferenceSession')
    @patch('components.debug.compare.msquickcmp.cmp_process.utils.logger')
    @patch('components.debug.compare.msquickcmp.cmp_process.sp.dynamic_divide_onnx')
    @patch('components.debug.compare.msquickcmp.cmp_process.sp.get_memory_size_by_soc_type')
    @patch('components.debug.compare.msquickcmp.cmp_process.sp.generate_single_op_dir')
    @patch('components.debug.compare.msquickcmp.cmp_process.sp.broken')
    @patch('components.debug.compare.msquickcmp.cmp_process.OnnxGraph')
    @patch('components.debug.compare.msquickcmp.cmp_process.CmpArgsAdapter')
    @patch('components.debug.compare.msquickcmp.cmp_process.analyser')
    def test_atc_conversion_mocked(self, mock_analyser, mock_cmp_args, mock_OnnxGraph, mock_sp_broken,
                                  mock_sp_gen_dir, mock_sp_get_mem, mock_sp_divide,
                                  mock_logger, mock_infersess, mock_rule,
                                  mock_al_input_completion, mock_al_find_npy,
                                  mock_al_create_bin, mock_ms_makedirs, mock_time_strftime,
                                  mock_shutil_rmtree, mock_utils_merge_csv,
                                  mock_run, mock_sp_find_all_csv, mock_sp_atc_conversion):

        class Args:
            model_path = "/fake/model.onnx"
            out_path = "/fake/output"
            device = "fake_device"
            cann_path = "/fake/cann"
        args = Args()
        input_shape = [1, 3, 224, 224]
        mock_og_instance = MagicMock()
        mock_OnnxGraph.parse.return_value = mock_og_instance
        mock_sp_broken.return_value = None
        mock_sp_gen_dir.return_value = "/fake/output/single_op_dir"
        mock_sp_get_mem.return_value = 1024

        mock_sp_divide.return_value = ["/fake/output/subonnx1.onnx", "/fake/output/subonnx2.onnx"]
        mock_sp_atc_conversion.return_value = None

        mock_session_instance = MagicMock()
        mock_session_instance.get_inputs.return_value = [
            MagicMock(name='input1', shape=[1, 3, 224, 224])
        ]
        mock_infersess.return_value = mock_session_instance
        mock_rule.input_file.return_value.check.return_value = None
        mock_al_input_completion.return_value = ['input1']
        mock_al_find_npy.return_value = ['/fake/output/dump_data/onnx/input1_data.npy']
        mock_al_create_bin.return_value = "/fake/output/bin_files"
        mock_ms_makedirs.return_value = None
        mock_time_strftime.return_value = "20250610123000"
        mock_run.return_value = "run_result"
        mock_sp_find_all_csv.side_effect = [
            ['/fake/output/single_op_dir/csv1.csv'],
            ['/fake/output/single_op_dir/csv2.csv']
        ]
        mock_utils_merge_csv.return_value = "/fake/output/single_op_dir/single_op_summary.csv"
        mock_analyser.Analyser.return_value = lambda: None

        single_op_compare(args, input_shape)
        expected_calls = [
            ("/fake/output/subonnx1.onnx", os.path.join(args.out_path, 'broken')),
            ("/fake/output/subonnx2.onnx", os.path.join(args.out_path, 'broken'))
        ]
        actual_calls = [call_args.args for call_args in mock_sp_atc_conversion.call_args_list]
        self.assertEqual(actual_calls, expected_calls)
        mock_sp_broken.assert_called_once()
        mock_sp_gen_dir.assert_called_once_with(args.out_path)
        mock_sp_get_mem.assert_called_once_with(args.device)
        mock_sp_divide.assert_called_once()
        mock_run.assert_called()
        mock_utils_merge_csv.assert_called_once()
        mock_analyser.Analyser.assert_called_once_with(mock_utils_merge_csv.return_value)
