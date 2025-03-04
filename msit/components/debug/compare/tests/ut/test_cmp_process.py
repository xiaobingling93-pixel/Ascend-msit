# Copyright (c) 2023-2024 Huawei Technologies Co., Ltd.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import shutil
import unittest
from unittest.mock import Mock

import pandas as pd
import openpyxl
import pytest

from msquickcmp.cmp_process import csv_sum, _process_is_npu_and_is_precision_error_ops, \
    _is_row_precision_error, _is_output_node, _get_model_output_node_name_list, \
    _find_previous_node

FAKE_CSV_PATH = "./test_resource/test_csv_sum"


@pytest.fixture(scope="function")
def generate_fake_path():

    os.mkdir(FAKE_CSV_PATH, 0o750)
    sub_folder_name = os.path.join(FAKE_CSV_PATH, "2023072009")
    os.mkdir(sub_folder_name, 0o750)
    os.mkdir('./test_resource/test_csv_sum/2023072009/images-2_3_638_640', 0o750)
    os.mkdir('./test_resource/test_csv_sum/2023072009/images-2_3_640_640', 0o750)

    df1 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df1.to_csv('./test_resource/test_csv_sum/2023072009/images-2_3_638_640/file1.csv', index=False)

    df2 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df2.to_csv('./test_resource/test_csv_sum/2023072009/images-2_3_640_640/file2.csv', index=False)

    with pd.ExcelWriter("./test_resource/test_csv_sum/expected_output.xlsx") as writer:
        df1.to_excel(writer, sheet_name='images-2_3_638_640', index=False)
        df2.to_excel(writer, sheet_name='images-2_3_640_640', index=False)

    yield sub_folder_name
    shutil.rmtree(FAKE_CSV_PATH)


def test_csv_sum_given_path_when_valid_then_pass(generate_fake_path):
    csv_sum(generate_fake_path)
    result_summary = openpyxl.load_workbook('./test_resource/test_csv_sum/2023072009/result_summary.xlsx')
    expected_output = openpyxl.load_workbook('./test_resource/test_csv_sum/expected_output.xlsx')

    sheets1 = result_summary.sheetnames
    sheets2 = expected_output.sheetnames

    assert len(sheets1) == len(sheets2)

    for sheet_name in sheets1:
        sheet1 = result_summary[sheet_name]
        sheet2 = expected_output[sheet_name]

        assert sheet1.max_row == sheet2.max_row
        assert sheet1.max_column == sheet2.max_column

        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                assert sheet1.cell(row=row, column=col).value == sheet2.cell(row=row, column=col).value


class TestProcessIsNpuAndIsPrecisionErrorOps(unittest.TestCase):

    def test_process_function(self):
        self.header = ["GroundTruth", "OpType", "CosineSimilarity", 
                       "RelativeEuclideanDistance", "KullbackLeiblerDivergence", 
                       "RootMeanSquareError", "MeanRelativeError"]
        self.rows = [
            [],
            ["*", "Add", "0.85", "0.12", "0.05", "0.04", "0.03"],  # NPU op, precision error
            ["*", "Add", "0.999", "0.005", "0.001", "0.04", "0.03"], # NPU op, no precision error
            # not NPU op, precision error
            ["/extractor/Softmax", "Add", "0.85", "0.12", "0.05", "0.04", "0.03"],  
            # not NPU op, no precision error
            ["/extractor/Softmax", "Add", "0.999", "0.005", "0.001", "0.04", "0.03"],
            # RelativeEuclideanDistance 
            ["/extractor/Softmax", "Add", "0.999", "0.06", "0.001", "0.04", "0.03"],
            # KullbackLeiblerDivergence
            ["/extractor/Reshape_3", "Add", "0.999", "0.005", "0.007", "0.04", "0.03"],
            # RootMeanSquareError
            ["/extractor/Softmax", "Add", "0.999", "0.005", "0.001", "1.1", "0.03"], 
            # MeanRelativeError
            ["/extractor/Concat_1", "Add", "0.999", "0.005", "0.001", "0.04", "10"] 
        ]
        node_output_name_list = ["/extractor/Reshape_3", "/extractor/Concat_1"]
        
        processed_rows = _process_is_npu_and_is_precision_error_ops(self.header, self.rows, node_output_name_list)

        self.assertIn('IsNpuOps', self.header)
        self.assertIn('IsOutputNode', self.header)
        self.assertIn('IsPrecisionError', self.header)

        expected_results = [
            ["YES", "NO", "YES"],  
            ["YES", "NO", "NO"],  
            ["NO", "NO", "YES"],  
            ["NO", "NO", "NO"],
            ["NO", "NO", "YES"],
            ["NO", "YES", "YES"],
            ["NO", "NO", "YES"],
            ["NO", "YES", "YES"]     
        ]
        
        for i, row in enumerate(processed_rows[1:]):
            self.assertEqual(row[-3:], expected_results[i])


class TestIsRowPrecisionError(unittest.TestCase):

    def test_below_thresholds(self):
        # All metrics are below the threshold
        result = _is_row_precision_error(0.95, 0.05, 0.04, 0.01, 0.005)
        self.assertTrue(result)

    def test_cosine_similarity_below_threshold(self):
        # only cosine similarity is below the threshold
        result = _is_row_precision_error(0.85, 0.05, 0.04, 0.01, 0.005)
        self.assertTrue(result)

    def test_relative_euclidean_distance_above_threshold(self):
        # only relative Euclidean distance above threshold
        result = _is_row_precision_error(0.95, 0.15, 0.04, 0.01, 0.005)
        self.assertTrue(result)

    def test_kullback_leibler_divergence_above_threshold(self):
        # only Kullback-Leibler above threshold
        result = _is_row_precision_error(0.95, 0.05, 0.06, 0.01, 0.005)
        self.assertTrue(result)

    def test_root_mean_square_error_above_threshold(self):
        # only RootMeanSquareError is above the threshold
        result = _is_row_precision_error(0.95, 0.05, 0.04, 0.03, 0.005)
        self.assertTrue(result)

    def test_mean_relative_error_above_threshold(self):
        # only mean relative error is above the threshold
        result = _is_row_precision_error(0.95, 0.05, 0.04, 0.01, 0.02)
        self.assertTrue(result)

    def test_multiple_errors(self):
        # Multiple indicators do not meet the requirements
        result = _is_row_precision_error(0.85, 0.15, 0.06, 0.03, 0.02)
        self.assertTrue(result)
        
        
class TestIsOutputNodeError(unittest.TestCase):
    
    def test__is_output_node(self):
        node_output_name_list = []
        result_false = _is_output_node("/extractor/Concat_1", node_output_name_list)
        self.assertFalse(result_false)
        
        node_output_name_list = ["/extractor/Reshape_3", "/extractor/Concat_1"]
        result_true = _is_output_node("/extractor/Reshape_3", node_output_name_list)
        self.assertTrue(result_true)
        
        result_false = _is_output_node("", node_output_name_list)
        self.assertFalse(result_false)


class MockSession:
    def get_outputs(self):
        return [MockNode(name="output_1", outputs=["net_output"]), MockNode(name="output_2", outputs=["net_output"])]

class MockGraph:
    def __init__(self, nodes):
        self.node = nodes

class MockNode:
    def __init__(self, name, outputs):
        self.name = name
        self.output = outputs


def test_find_previous_node():
    # 创建模拟的节点
    node1 = MockNode(name="node1", outputs=["output_1"])
    node2 = MockNode(name="node2", outputs=["output_2"])
    graph = MockGraph(nodes=[node1, node2])

    # 测试找到前一个节点
    assert _find_previous_node(graph, "output_1") == "node1"
    assert _find_previous_node(graph, "output_2") == "node2"

    # 测试未找到前一个节点
    assert _find_previous_node(graph, "output_3") is None


def test_get_model_output_node_name_list():
    # 创建模拟的会话和模型
    session = MockSession()
    node1 = MockNode(name="node1", outputs=["output_1"])
    node2 = MockNode(name="node2", outputs=["output_2"])
    origin_model = Mock(graph=MockGraph(nodes=[node1, node2]))
    # 测试正常情况
    result = _get_model_output_node_name_list(session, origin_model)
    assert result == ["node1", "node2"], f"Expected ['node1', 'node2'], but got {result}"

    # 测试找不到前一个节点的情况
    node3 = MockNode(name="node3", outputs=["output_3"])
    origin_model.graph.node = [node3]  # 修改模型节点
    result = _get_model_output_node_name_list(session, origin_model)
    assert result is None, f"Expected None, but got {result}"