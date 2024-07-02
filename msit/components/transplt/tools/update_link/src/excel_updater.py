import os

import openpyxl
import pandas as pd
from selenium.webdriver.common.by import By

from .logger import logger
from .util import open_excel
from .web_crawler import ASCEND_URL_PREFIX


class ConfigExcelUpdater:
    def __init__(self, crawler, src_dir, new_api_link_file):
        self._src_dir = src_dir
        self.new_api_link_file = new_api_link_file
        self.crawler = crawler
        self.origin_api_link_dict = {}

    @staticmethod
    def _read_excel(file_path=""):
        excel = pd.ExcelFile(file_path, engine='openpyxl')
        sheets = excel.sheet_names

        df_dict = dict()
        for sheet in sheets:
            df = excel.parse(sheet)
            df_dict[sheet] = df
        return df_dict

    @staticmethod
    def _write_excel_keep_format(df_dict, path='output.xlsx'):
        excel = open_excel(path)
        for key, df in df_dict.items():
            if "APIMap" not in key:
                continue
            logger.info(f"write sheet {key} to excel: {path}")
            df.to_excel(excel, sheet_name=key, index=False)
        excel.close()

    def _crawl_new_ascend_api_link(self, new_api_link_dict, api_link):
        if not api_link.startswith(ASCEND_URL_PREFIX):
            return api_link

        logger.info(f"crawling url: {api_link}")
        success, is_throttled = self.crawler.open_ascend_url(api_link)

        if is_throttled:
            logger.error(f"failed to crawl url {api_link}!!")
            return api_link

        api_name = self.crawler.driver.find_element(By.CLASS_NAME, "topic-title").text.strip()

        fixed_api_name = api_name
        tmp_class_name = self.crawler.try_get_api_class_name()
        if len(tmp_class_name) > 0:
            fixed_api_name = f"{tmp_class_name}::{api_name}"

        if fixed_api_name in new_api_link_dict:
            api_link = new_api_link_dict.get(fixed_api_name)

        return api_link

    def _read_all_origin_api_links(self):
        for path, _, file_list in os.walk(self._src_dir):
            for f in file_list:
                if not f.endswith("xlsx"):
                    continue

                file_path = os.path.join(path, f)
                logger.info(f"pre processing file: {file_path}")
                df_dict = self._read_excel(file_path)

                self._parse_api_links_of_one_excel(df_dict)

    def _parse_api_links_of_one_excel(self, df_dict):
        for sheet_name, df in df_dict.items():
            if not sheet_name.endswith("APIMap"):
                continue
            datas = df.loc[:, ["AscendAPI", "AscendAPILink"]].values.tolist()
            for index, values in enumerate(datas):
                api_name = values[0]
                api_link = values[1]
                if not type(api_link) is str or "https" not in api_link:
                    continue
                self.origin_api_link_dict[api_name] = api_link

    def _update_excel_api_link_text(self, df_dict, file_path, new_api_link_dict):
        """
        更新excel中的API文档链接
        :param df_dict:
        :param file_path:
        :param new_api_link_dict:
        :return:
        """
        for sheet_name, df in df_dict.items():
            if "APIMap" not in sheet_name:
                continue
            datas = df.loc[:, ["AscendAPI", "AscendAPILink"]].values.tolist()
            for index, values in enumerate(datas):
                api_name = values[0]
                api_link = values[1]

                if type(api_name) is not str:
                    continue

                new_api_link = api_link
                if api_name in new_api_link_dict:
                    new_api_link = new_api_link_dict.get(api_name)
                elif api_name.split(".")[0] in new_api_link_dict:
                    # 部分execl中的api名字是结构体的成员变量，以"."分割，尝试获取正确的名字
                    new_api_link = new_api_link_dict.get(api_name.split(".")[0])
                elif type(api_link) is str and api_link.startswith("https"):
                    # 部分execl中的api名字是结构体的成员变量，在new_api_link_dict中无法找到，需要爬取网页并从网页中获取类的名字
                    new_api_link = self._crawl_new_ascend_api_link(new_api_link_dict, api_link)
                elif api_name in self.origin_api_link_dict:
                    # AscendAPILink 为空， 尝试从其它excel中读取已有的api link，并更新到最新
                    tmp_api_link = self.origin_api_link_dict[api_name]
                    new_api_link = self._crawl_new_ascend_api_link(new_api_link_dict, tmp_api_link)
                else:
                    logger.error(f"unable to update link of api {api_name} in excel {file_path}")

                logger.debug("update link to:", new_api_link)
                df_dict[sheet_name].loc[index:index, "AscendAPILink"] = new_api_link

        self._write_excel_keep_format(df_dict, file_path)

    def _pre_process_excels(self):
        for path, _, file_list in os.walk(self._src_dir):
            for f in file_list:
                if not f.endswith("xlsx"):
                    continue

                file_path = os.path.join(path, f)

                logger.info(f"preprocess excel: {file_path}")
                self._update_execl_api_text_according_to_hyperlink(file_path)

    def _read_new_api_links(self):
        df_dict = self._read_excel(self.new_api_link_file)
        api_link_dict = {}
        for df in df_dict.values():
            datas = df.loc[:, ["URL", "Fixed Function Name"]].values.tolist()
            api_link_dict.update({value[1]: value[0] for value in datas})
        return api_link_dict

    def _update_excel_api_link_hyperlink(self, file_path=None):
        """
        更新excel中API文档cell数据的超链接
        :param file_path:
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            self._update_hyperlink_of_one_cell(wb[sheet_name])
        wb.save(file_path)

    def _update_execl_api_text_according_to_hyperlink(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            if not sheet_name.endswith("APIMap"):
                continue
            logger.info(f"preprocess sheet: {sheet_name}")
            self._update_text_of_one_cell_according_to_hyperlink(wb[sheet_name])
        wb.save(file_path)

    @staticmethod
    def _update_text_of_one_cell_according_to_hyperlink(ws):
        """
        根据cell的超链接更新cell的链接文本
        :param ws:
        :return:
        """
        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                text = ws.cell(row=i, column=j).value
                if type(text) is str and text.startswith("https"):
                    continue
                if not ws.cell(row=i, column=j).hyperlink:
                    continue
                link = ws.cell(row=i, column=j).hyperlink.target
                if type(link) is str and link.startswith("https"):
                    ws.cell(row=i, column=j).value = link

    @staticmethod
    def _update_hyperlink_of_one_cell(ws):
        """
        根据cell的链接文本更新cell的超链接
        :param ws:
        :return:
        """
        for i in range(1, ws.max_row):
            for j in range(1, ws.max_column):
                text = ws.cell(row=i, column=j).value
                if type(text) is str and text.startswith("https"):
                    ws.cell(row=i, column=j).hyperlink = text

    def update_excels(self):
        self._pre_process_excels()

        self._read_all_origin_api_links()

        new_api_link_dict = self._read_new_api_links()

        for path, _, file_list in os.walk(self._src_dir):
            for f in file_list:
                if not f.endswith("xlsx"):
                    continue

                file_path = os.path.join(path, f)
                logger.info(f"processing file: {file_path}")
                df_dict = self._read_excel(file_path)

                try:
                    self._update_excel_api_link_text(df_dict, file_path, new_api_link_dict)
                    self._update_excel_api_link_hyperlink(file_path)
                except Exception as err:
                    logger.error(f"failed to update_excel {file_path}, error is: {err}")

        self.crawler.close()
