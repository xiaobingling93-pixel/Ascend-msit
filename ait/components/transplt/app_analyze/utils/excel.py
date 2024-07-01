# Copyright (c) 2023-2024 Huawei Technologies Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
import argparse
import os
import re

import pandas as pd
import openpyxl
from xlsxwriter.workbook import Workbook

from components.utils.file_open_check import PERMISSION_NORMAL, FileStat, OpenException
from app_analyze.utils.log_util import logger


def get_url(cell):
    return cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value


def update_hyperlink(path, sheet, hyperlink_cols, df=None, row_header=1):
    """更新df中超链接的text为url。
    Args:
        path: Excel路径。
        sheet: 工作表名。
        hyperlink_cols: 包含超链接的列。
        df: 已读取的Excel生成的DataFrame对象。
        row_header: 行表头的行号，超链接从下一行开始。
    """
    if df is None:
        df = pd.read_excel(path, sheet)
    ws = openpyxl.load_workbook(path)[sheet]
    # ws的坐标从1开始
    for col_name in hyperlink_cols:
        row = row_header + 1
        if col_name not in df.columns:
            continue
        col = list(df.columns).index(col_name) + 1
        df[col_name] = [get_url(ws.cell(row=row + i, column=col)) for i in range(len(df))]
    return df


def read_excel(path="", hyperlink_cols=None):
    try:
        file_stat = FileStat(path)
    except Exception as ee:
        raise argparse.ArgumentTypeError(f"input excel path:{path} is illegal. Please check.") from ee
    else:
        if not file_stat.is_basically_legal('read'):
            raise argparse.ArgumentTypeError(f"input excel path:{path} is illegal. Please check.")
    # 读取Excel文件
    excel = pd.ExcelFile(path)
    # 获取所有Sheet的名称
    sheets = excel.sheet_names

    api_dfs = dict()
    # 遍历每个Sheet
    for sheet in sheets:
        if sheet.startswith('Sheet'):  # 无效Sheet
            continue
        # 读取当前Sheet的数据为DataFrame对象
        df = excel.parse(sheet)
        # 获取列名列表
        column_names = df.columns.tolist()
        # 打印当前Sheet的名称和列名
        logger.debug(f'{sheet} {column_names}')
        df = update_hyperlink(path, sheet, hyperlink_cols, df)
        api_dfs[sheet] = df
    return api_dfs


# mode: df or workbook
def write_excel(df_dict, path='output.xlsx'):
    # 创建一个 Excel 文件
    excel = pd.ExcelWriter(path)
    for key, df in df_dict.items():
        key = re.sub(r"\W", ".", key)[-31:]  # 最大支持31个字符
        df.to_excel(excel, sheet_name=key, index=False)
    # 保存 Excel 文件
    excel.save()
    # set permission to 640
    os.chmod(path, PERMISSION_NORMAL)


def df2xlsx(df_dict, fmt_dict, path='output.xlsx'):
    workbook = Workbook(path)
    for key, df in df_dict.items():
        fmt = fmt_dict[key]
        key = key.replace('/', '.')[-31:]  # 最大支持31个字符
        fmt(key, df, workbook)
    workbook.close()
    # set permission to 640
    os.chmod(path, PERMISSION_NORMAL)
